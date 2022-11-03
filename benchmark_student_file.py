import openpyxl
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

#student_file = 'students.xlsx'
#clever_file = 'clever.xlsx'

import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()

print("Select student file")
student_file = filedialog.askopenfilename()
print("Student Synergy File:",student_file)
print("Select clever file")
clever_file = filedialog.askopenfilename()
print("Clever File:",student_file)

#convert to XLSX format
clever_file = pd.read_csv (clever_file)
clever_file.to_excel (r'clever.xlsx', index = None, header=True)
clever_file = 'clever.xlsx'

# Change the Kindergarten Grade code from KF or 25 to K
def fixKindergartenGradeLevel(file):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    for row in range (1,900): #stops at row 900
        for col in range (10,11): # column J
            char = get_column_letter(col)
            if ws[char + str(row)].value == "KF" or ws[char + str(row)].value == "25":
                ws[char + str(row)].value = 'K'
    wb.save(file)
            

# Alter the Clever file 
def formatCleverFile(file):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    ws['E1'].value = "Student's SIS Id"
    wb.save(file)

def mergeSheets(student_file,clever_file):
    #merging data frames with merge function in pandas
    # load in the various tables from an excel document
    studentData = pd.read_excel(student_file,sheet_name='QRY801')
    #cleverData = pd.read_excel(clever_file,sheet_name='clever') # needs modification
    cleverData = pd.read_excel(clever_file,sheet_name='Sheet1')
    merge = studentData.merge(cleverData, on="Student's SIS Id", how='left')
    #print(merge)
    #merge.to_csv('merge.csv',index=False)
    merge.to_excel('merge.xlsx',index=False)

# Format the columns of the newly merged sheet
def formatSheet(file):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    #print(ws)
    ws.delete_cols(14)
    ws.delete_cols(15)
    ws.delete_cols(14)
    ws.delete_cols(13)
    for row in range (1,2): #stops at row 10
     for col in range (11,15): #columns 1 -4
         char = get_column_letter(col)
         #print(ws[char + str(row)].value)
    
    #Move Column M to Column K
    for row in range (1,900): #stops at row 10
     for col in range (11,12): #columns 1 -4
        char = get_column_letter(col)
       
        ws[char + str(row)].value = ws[get_column_letter(13) + str(row)].value
        #print(ws[char + str(row)].value)
    # change colmun header
    ws['K1'].value = 'SSO Id'
    ws.delete_cols(13) 
    wb.save(file)


#Main Program 

if __name__ == "__main__":
    print("Starting...")
    fixKindergartenGradeLevel(student_file)
    formatCleverFile(clever_file)
    mergeSheets(student_file,clever_file)
    formatSheet('merge.xlsx')
    cleverData = pd.read_excel('merge.xlsx',sheet_name='Sheet1')
    cleverData.to_csv("students_IMPORT.csv",index=False)
    print("finished")
